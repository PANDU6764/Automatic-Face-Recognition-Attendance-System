import time
import cv2
import os
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook, load_workbook

from app.settings import (
    DB_PATH,
    MIN_CONFIDENCE, L2_THRESHOLD,
    CAMERA_INDEX, FRAME_W, FRAME_H
)

from src.utils.camera import open_camera
from src.detection.mtcnn_detector import load_mtcnn, detect_faces
from src.recognition.embedder import Embedder
from src.recognition.matcher import match_face
from src.storage.sqlite_db import connect, init_db, load_persons


# ---------------- STORAGE PATHS ----------------
BASE_DIR = "data/snapshots"
IMG_DIR = os.path.join(BASE_DIR, "images")
EXCEL_PATH = os.path.join(BASE_DIR, "attendance.xlsx")

os.makedirs(IMG_DIR, exist_ok=True)


# ---------------- EXCEL FUNCTIONS ----------------
def init_excel():
    if os.path.exists(EXCEL_PATH):
        return

    wb = Workbook()
    ws_p = wb.active
    ws_p.title = "Present"
    ws_p.append(["Date", "Time", "Roll No", "Name", "Image Path"])

    ws_a = wb.create_sheet("Absent")
    ws_a.append(["Date", "Roll No", "Name"])

    wb.save(EXCEL_PATH)


def save_present(date_str, time_str, pid, name, img_path):
    wb = load_workbook(EXCEL_PATH)
    wb["Present"].append([date_str, time_str, pid, name, img_path])
    wb.save(EXCEL_PATH)


def save_absentees(date_str, all_persons, present_ids):
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Absent"]

    for pid, name, _ in all_persons:
        if pid not in present_ids:
            ws.append([date_str, pid, name])

    wb.save(EXCEL_PATH)


# ---------------- REAL-TIME ATTENDANCE ----------------
def run_realtime(session_seconds=30):
    """
    Stable real-time attendance
    - EXACT 30 second session
    - Laptop camera friendly
    - One person → one attendance
    """

    init_excel()

    detector = load_mtcnn()
    embedder = Embedder()

    conn = connect(DB_PATH)
    init_db(conn)
    persons = load_persons(conn)

    if not persons:
        raise RuntimeError("❌ No enrolled persons found.")

    cap = open_camera(CAMERA_INDEX, FRAME_W, FRAME_H)

    marked_ids = set()
    identity_buffer = defaultdict(int)

    STABILITY_FRAMES = 3   # 🔑 RELAXED & SAFE

    today = datetime.now().strftime("%Y-%m-%d")
    start_time = time.time()

    print("🎥 Attendance started (30 seconds)...")

    try:
        while True:
            elapsed = time.time() - start_time
            if elapsed >= session_seconds:
                print("⏱ Attendance session ended automatically")
                break

            ok, frame = cap.read()
            if not ok:
                break

            dets = detect_faces(detector, frame, MIN_CONFIDENCE)

            # ---- FACE PROCESSING ----
            if len(dets) >= 1:
                d = dets[0]  # strongest face only
                x, y, w, h = d["box"]

                emb = embedder.embed_from_detection(frame, d)
                pid, name, dist = match_face(emb, persons, L2_THRESHOLD)

                if pid:
                    identity_buffer[pid] += 1

                    if identity_buffer[pid] >= STABILITY_FRAMES and pid not in marked_ids:
                        now = datetime.now()
                        time_str = now.strftime("%H:%M:%S")

                        img_name = f"{pid}_{now.strftime('%Y%m%d_%H%M%S')}.jpg"
                        img_path = os.path.join(IMG_DIR, img_name)

                        face_crop = frame[y:y+h, x:x+w]
                        cv2.imwrite(img_path, face_crop)

                        save_present(today, time_str, pid, name, img_path)
                        marked_ids.add(pid)

                        print(f"✅ Marked Present: {pid}")
                        identity_buffer.clear()
                else:
                    identity_buffer.clear()

                label = f"{name if pid else 'Unknown'}  L2={dist:.2f}"
                color = (0, 255, 0) if pid else (0, 0, 255)

                cv2.rectangle(frame, (x, y), (x+w, y+h), color, 2)
                cv2.putText(
                    frame, label, (x, y - 10),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.7, color, 2
                )

            # ---- COUNTDOWN TIMER ----
            remaining = max(0, int(session_seconds - elapsed))
            cv2.putText(
                frame,
                f"Time Left: {remaining}s",
                (20, 40),
                cv2.FONT_HERSHEY_SIMPLEX,
                1,
                (0, 0, 255),
                2
            )

            cv2.imshow("AutoFace Attendance (30s)", frame)
            cv2.waitKey(1)

    finally:
        cap.release()
        cv2.destroyAllWindows()

        save_absentees(today, persons, marked_ids)
        conn.close()

        print("\n📊 Attendance completed")
        print(f"👥 Present: {len(marked_ids)}")
        print(f"🚫 Absent : {len(persons) - len(marked_ids)}")
